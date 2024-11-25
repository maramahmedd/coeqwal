import openpyxl
import re

def get_xl_sheetnames(xlfn):
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    return(sheet_names)

def read_from_excel(xlfn, tabname, topleft, bottomright,hdr=True,dtypes=[]):
    # open excel file
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    inputs_sheet = wb.get_sheet_by_name(tabname)
    topsplit = re.split('(\d+)', topleft)
    botsplit = re.split('(\d+)', bottomright)
    
    if hdr:
        hdr_block = inputs_sheet[topleft:(botsplit[0]+topsplit[1])]
        hdr_values = [[str(v.value) for v in v1] for v1 in hdr_block][0]                        
        data_block = inputs_sheet[topsplit[0]+str(int(topsplit[1])+1):bottomright]
    else:
        data_block = inputs_sheet[topleft:bottomright]   
        hdr_values = None
    data_list = []
    # loop over months
    for i,row in enumerate(data_block):
        # loop over columns
        tmp  =[]
        if row[0].value=='null':
            pass
        else:
            if not dtypes: # list of dtypes is empty
                for j,cell in enumerate(row[:]):
                    tmp.append(str(cell.value))
            elif len(dtypes) != len(row): 
                print("Wrong number of dtypes provided - returning values as strings")
            else:
                for j,cell in enumerate(row[:]):
                    dtypi = dtypes[j]
                    if dtypi[0:2]=='dt':
                        parsetxt=dtypi[2:]
                        if type(cell.value) is not dt.datetime:  #check if it's already parsed as a datetime
                            tmp.append(dt.datetime.strptime(cell.value,parsetxt))
                        else:
                            tmp.append((cell.value))
                    elif dtypi[0:5]=='float':
                        tmp.append(float(cell.value))
                    else:
                        tmp.append(str(cell.value))
            data_list.append(tmp)
    return([hdr_values, data_list])