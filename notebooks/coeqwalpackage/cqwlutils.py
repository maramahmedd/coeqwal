import openpyxl
import re
import os

def get_xl_sheetnames(xlfn):
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    return(sheet_names)

def read_from_excel(xlfn, tabname, topleft, bottomright,hdr=True,dtypes=[]):
    # open excel file
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    inputs_sheet = wb.get_sheet_by_name(tabname)
    topsplit = re.split('(\d+)', topleft)
    botsplit = re.split('(\d+)', bottomright)
    
    if hdr:
        hdr_block = inputs_sheet[topleft:(botsplit[0]+topsplit[1])]
        hdr_values = [[str(v.value) for v in v1] for v1 in hdr_block][0]                        
        data_block = inputs_sheet[topsplit[0]+str(int(topsplit[1])+1):bottomright]
    else:
        data_block = inputs_sheet[topleft:bottomright]   
        hdr_values = None
    data_list = []
    # loop over months
    for i,row in enumerate(data_block):
        # loop over columns
        tmp  =[]
        if row[0].value=='null':
            pass
        else:
            if not dtypes: # list of dtypes is empty
                for j,cell in enumerate(row[:]):
                    tmp.append(str(cell.value))
            elif len(dtypes) != len(row): 
                print("Wrong number of dtypes provided - returning values as strings")
            else:
                for j,cell in enumerate(row[:]):
                    dtypi = dtypes[j]
                    if dtypi[0:2]=='dt':
                        parsetxt=dtypi[2:]
                        if type(cell.value) is not dt.datetime:  #check if it's already parsed as a datetime
                            tmp.append(dt.datetime.strptime(cell.value,parsetxt))
                        else:
                            tmp.append((cell.value))
                    elif dtypi[0:5]=='float':
                        tmp.append(float(cell.value))
                    else:
                        tmp.append(str(cell.value))
            data_list.append(tmp)
    return([hdr_values, data_list])


def read_init_file(CtrlFile, CtrlTab):

    # Hard-coded file structure
    ScenarioDirInd = 'B2'
    ScenarioListFileInd = 'B3'
    ScenarioListTabInd = 'B4'
    IndexMinInd = 'C5' # first scenario index
    IndexMaxInd = 'D5' # last scenario index
    ScenariosNameMinInd = 'C6' # first dss dir index
    ScenariosNameMaxInd = 'D6' # last dss dir index
    ScenariosDirMinInd = 'C7' # first dss dir index
    ScenariosDirMaxInd = 'D7' # last dss dir index
    DVDssPathMinInd = 'C8' # first dss path index
    DVDssPathMaxInd = 'D8' # last dss path index
    SVDssPathMinInd = 'C9' # first dss path index
    SVDssPathMaxInd = 'D9' # last dss path index
    StartMinInd = 'C10' # first start date index
    StartMaxInd = 'D10' # last start date index
    EndMinInd = 'C11' # First end date index
    EndMaxInd = 'D11' # Last end date index
    GroupDataDirInd = 'B12' # Group data extraction directory
    VarFileNameInd = 'B13' # Variable list file name
    VarFileTabInd = 'B14' # Variable list file tab
    VarMinInd = 'C15' # top left of variable name block
    VarMaxInd = 'D15' # bottom right of variable name block
    ExtractionDirInd = 'B16'
    DemDelDirInd = 'B17'
    ModelFilesDirInd = 'B18'
    ModelSubDirInd = 'B19'
    DemandsFileInd = 'B20'
    DemandsTabInd = 'B21'
    DemMinInd = 'C22' # top left of demand name block
    DemMaxInd = 'D22' # bottom right of demand name block

    InflowDirInd = 'B23'
    InflowFileInd = 'B24'
    InflowTabInd = 'B25'
    InflowMinInd = 'C26' # top left of demand name block
    InflowMaxInd = 'D26' # bottom right of demand name block

    # Control File Example
    #Item	Name or description	Upper Left Cell	Lower Right Cell
    #Scenarios Directory	../../CalSim3_Model_Runs/Scenarios		
    #Scenario Listings File	coeqwal_cs3_scenario_listing_v2.xlsx		
    #Scenario Listings Tab	scenario_list		
    #Scenario Indices	Scenario identifiers	A1	A11
    #Scenario Directory Indices	Scenario directory names	C1	C11
    #DSS Path Indices	Dss path names	G1	G11
    #Start Date Indices	Start dates	H1	H11
    #Start Date Indices	End dates	I1	I11
    #Group Extraction Directory	Group_Data_Extraction		
    #Variables Listing File	trend_report_variables_v3.xlsx		
    #Variables List Tab	Variables List		E177
    #Variables List Indices	Variables List Block	D8	E177
    #Data Extraction Dir	Data_Extraction		
    #Model Files Dir	Model_Files		
    #Model Files SubDir	ModelFiles		

    # Read directory structure and contol file name
    Hdr, ScenarioDir = read_from_excel(CtrlFile, CtrlTab, ScenarioDirInd, ScenarioDirInd, hdr=False) # Scenarios directory (../../CalSim3_Model_Runs/Scenarios in the current structure)
    ScenarioDir = ScenarioDir[0][0]
    Hdr, ScenarioListFile = read_from_excel(CtrlFile, CtrlTab, ScenarioListFileInd, ScenarioListFileInd, hdr=False) # DSS file names Excel workbook
    ScenarioListFile = ScenarioListFile[0][0]
    ScenarioListPath = os.path.join(ScenarioDir,ScenarioListFile) # path to DSS file names Excel workbook
    
    # Read file names and ranges for DSS and vars from control file
    Hdr, ScenarioListTab = read_from_excel(CtrlFile, CtrlTab, ScenarioListTabInd, ScenarioListTabInd, hdr=False) # DSS file names Excel workbook Tab
    ScenarioListTab = ScenarioListTab[0][0]
    Hdr, IndexMin = read_from_excel(CtrlFile, CtrlTab, IndexMinInd, IndexMinInd, hdr=False) # Scenario Index Name UL
    IndexMin = IndexMin[0][0]
    Hdr, IndexMax = read_from_excel(CtrlFile, CtrlTab, IndexMaxInd, IndexMaxInd, hdr=False) # Scenario Index Name LR
    IndexMax = IndexMax[0][0]
    Hdr, NameMin = read_from_excel(CtrlFile, CtrlTab, ScenariosNameMinInd, ScenariosNameMinInd, hdr=False) # Scenario Name UL
    NameMin = NameMin[0][0]
    Hdr, NameMax = read_from_excel(CtrlFile, CtrlTab, ScenariosNameMaxInd, ScenariosNameMaxInd, hdr=False) # Scenario Name UL
    NameMax = NameMax[0][0]
    Hdr, DirMin = read_from_excel(CtrlFile, CtrlTab, ScenariosDirMinInd, ScenariosDirMinInd, hdr=False) # Scenario Dir Name UL
    DirMin = DirMin[0][0]
    Hdr, DirMax = read_from_excel(CtrlFile, CtrlTab, ScenariosDirMaxInd, ScenariosDirMaxInd, hdr=False) # Scenario Dir Name UL
    DirMax = DirMax[0][0]
    Hdr, DVDssMin = read_from_excel(CtrlFile, CtrlTab, DVDssPathMinInd, DVDssPathMinInd, hdr=False) # DSS Path Name UL
    DVDssMin = DVDssMin[0][0]
    Hdr, DVDssMax = read_from_excel(CtrlFile, CtrlTab, DVDssPathMaxInd, DVDssPathMaxInd, hdr=False) # DSS Path Name LR
    DVDssMax = DVDssMax[0][0]
    Hdr, SVDssMin = read_from_excel(CtrlFile, CtrlTab, SVDssPathMinInd, SVDssPathMinInd, hdr=False) # DSS Path Name UL
    SVDssMin = SVDssMin[0][0]
    Hdr, SVDssMax = read_from_excel(CtrlFile, CtrlTab, SVDssPathMaxInd, SVDssPathMaxInd, hdr=False) # DSS Path Name LR
    SVDssMax = SVDssMax[0][0]
    Hdr, StartMin = read_from_excel(CtrlFile, CtrlTab, StartMinInd, StartMinInd, hdr=False) # Start Date UL
    StartMin = StartMin[0][0]
    Hdr, StartMax = read_from_excel(CtrlFile, CtrlTab, StartMaxInd, StartMaxInd, hdr=False) # Start Date LR
    StartMax = StartMax[0][0]
    Hdr, EndMin = read_from_excel(CtrlFile, CtrlTab, EndMinInd, EndMinInd, hdr=False) # Start Date UL
    EndMin = EndMin[0][0]
    Hdr, EndMax = read_from_excel(CtrlFile, CtrlTab, EndMaxInd, EndMaxInd, hdr=False) # Start Date LR
    EndMax = EndMax[0][0]
    Hdr, GroupDataDirName = read_from_excel(CtrlFile, CtrlTab, GroupDataDirInd, GroupDataDirInd, hdr=False) # directory name for group data extraction (Group_Data_Extraction in current structure)
    GroupDataDirName = GroupDataDirName[0][0]
    GroupDataDirPath = os.path.join(ScenarioDir, GroupDataDirName) # group data extraction directory (../../CalSim3_Model_Runs/Scenarios/Group_Data_Extraction in the current structure)
    #print(GroupDataDirPath)
    Hdr, VarListFileName = read_from_excel(CtrlFile, CtrlTab, VarFileNameInd, VarFileNameInd, hdr=False) # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    VarListFileName = VarListFileName[0][0]
    Hdr, VarListTab = read_from_excel(CtrlFile, CtrlTab, VarFileTabInd, VarFileTabInd, hdr=False) # tab for variable listing (TrendReportVars_CS3 in current structure)
    VarListTab = VarListTab[0][0]
    Hdr, VarMin = read_from_excel(CtrlFile, CtrlTab, VarMinInd, VarMinInd, hdr=False) # variable listing UL
    VarMin = VarMin[0][0]
    Hdr, VarMax = read_from_excel(CtrlFile, CtrlTab, VarMaxInd, VarMaxInd, hdr=False) # variable listing LR
    VarMax = VarMax[0][0]
    Hdr, ExtractionDir = read_from_excel(CtrlFile, CtrlTab, ExtractionDirInd, ExtractionDirInd, hdr=False) #  Var extraction Dir Name
    ExtractionDir = ExtractionDir[0][0]
    Hdr, DemandDeliveryDir = read_from_excel(CtrlFile, CtrlTab, DemDelDirInd, DemDelDirInd, hdr=False) #  Var extraction Dir Name
    DemandDeliveryDir = DemandDeliveryDir[0][0]
    Hdr, ModelFilesDir = read_from_excel(CtrlFile, CtrlTab, ModelFilesDirInd, ModelFilesDirInd, hdr=False) #  Var extraction Dir Name
    ModelFilesDir = ModelFilesDir[0][0]
    Hdr, ModelSubDir = read_from_excel(CtrlFile, CtrlTab, ModelSubDirInd, ModelSubDirInd, hdr=False) #  Var extraction SubDir Name
    ModelSubDir = ModelSubDir[0][0]
    Hdr, DemandFileName = read_from_excel(CtrlFile, CtrlTab, DemandsFileInd, DemandsFileInd, hdr=False) # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    DemandFileName = DemandFileName[0][0]
    Hdr, DemandFileTab = read_from_excel(CtrlFile, CtrlTab, DemandsTabInd, DemandsTabInd, hdr=False) # tab for variable listing (TrendReportVars_CS3 in current structure)
    DemandFileTab = DemandFileTab[0][0]
    Hdr, DemMin = read_from_excel(CtrlFile, CtrlTab, DemMinInd, DemMinInd, hdr=False) # variable listing UL
    DemMin = DemMin[0][0]
    Hdr, DemMax = read_from_excel(CtrlFile, CtrlTab, DemMaxInd, DemMaxInd, hdr=False) # variable listing LR
    DemMax = DemMax[0][0]
    
    Hdr, InflowDir = read_from_excel(CtrlFile, CtrlTab, InflowDirInd, InflowDirInd, hdr=False) #  Var extraction Dir Name
    InflowDir = InflowDir[0][0]
    Hdr, InflowFileName = read_from_excel(CtrlFile, CtrlTab, InflowFileInd, InflowFileInd, hdr=False) # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    InflowFileName = InflowFileName[0][0]
    Hdr, InflowFileTab = read_from_excel(CtrlFile, CtrlTab, InflowTabInd, InflowTabInd, hdr=False) # tab for variable listing (TrendReportVars_CS3 in current structure)
    InflowFileTab = InflowFileTab[0][0]
    Hdr, InflowMin = read_from_excel(CtrlFile, CtrlTab, InflowMinInd, InflowMinInd, hdr=False) # variable listing UL
    InflowMin = InflowMin[0][0]
    Hdr, InflowMax = read_from_excel(CtrlFile, CtrlTab, InflowMaxInd, InflowMaxInd, hdr=False) # variable listing LR
    InflowMax = InflowMax[0][0]

    # Construct file and directory names
    # File and directory names
    ScenarioListFileCsv = ScenarioListFile.replace(".xlsx", ".csv")
    DVDssNamesOut = 'DVDssNamesFrom_' + ScenarioListFileCsv # output DSS names CSV
    SVDssNamesOut = 'SVDssNamesFrom_' + ScenarioListFileCsv # output DSS names CSV
    ScenarioIndicesOut = 'IndicesFrom_' + ScenarioListFileCsv # output DSS indices CSV
    DssDirsOut = 'DirNamesFrom_' + ScenarioListFileCsv # output directory names CSV
    DVDssNamesOutPath = os.path.join(GroupDataDirPath, DVDssNamesOut) # output DSS names CSV path
    SVDssNamesOutPath = os.path.join(GroupDataDirPath, SVDssNamesOut) # output DSS names CSV path
    ScenarioIndicesOutPath = os.path.join(GroupDataDirPath, ScenarioIndicesOut) # output DSS index names CSV path
    DssDirsOutPath = os.path.join(GroupDataDirPath, DssDirsOut) # output DSS dir names CSV path

    # list of relevant variables file, tab, and range (B & C parts)
    VarListName = os.path.splitext(VarListFileName)[0] # variable names file without extension
    VarListExt = os.path.splitext(VarListFileName)[1] # variable names file extension
    VarListFile = VarListName + VarListExt # full file name
    VarListFileCsv = VarListFile.replace(".xlsx", ".csv")
    VarListPath = os.path.join(ScenarioDir, VarListFile)
    DemandFilePath = os.path.join(ScenarioDir, DemandFileName)
    VarOut = 'VarsFrom_' + VarListFileCsv # output compund variable names CSV
    VarOutPath = os.path.join(GroupDataDirPath, VarOut)
    DataOut = 'DataFrom_' + VarListFileCsv # file name for multi-study output CSV
    DataOutPath = os.path.join(GroupDataDirPath, DataOut) # file name for multi-study output CSV path
    ConvertDataOut = 'ConvertDataFrom_' + VarListFileCsv # file name for multi-study output CSV
    ConvertDataOutPath = os.path.join(GroupDataDirPath, ConvertDataOut) # file name for multi-study output CSV path
    ExtractionSubDir = 'Variables_From_' + VarListName + '_' + VarListTab
    ExtractionSubPath = os.path.join(ExtractionDir, ExtractionSubDir)
    DemandDeliverySubPath = os.path.join(ExtractionDir, DemandDeliveryDir)
    ModelSubPath = os.path.join('Model_Files','DSS','output')
    InflowOutSubPath = os.path.join(ExtractionDir, InflowDir)
    InflowFilePath = os.path.join(ScenarioDir, InflowFileName)

    # debug print
    # print(ScenarioListFile)    
    # print(ScenarioListTab)    
    # print(ScenariosistPath)
    # print(DssNamesOutPath)
    # print(ScenarioIndicesOutPath)
    # print(DssDirsOutPath)
    # print(VarListPath)
    # print(VarOutPath)
    # print(DataOutPath)
    # print(ExtractionSubPath)
    # print(ModelSubPath)
    # print(GroupDataDirPath)
 
    # return info
    return ScenarioListFile, ScenarioListTab, ScenarioListPath, DVDssNamesOutPath, SVDssNamesOutPath, ScenarioIndicesOutPath, DssDirsOutPath, VarListPath, VarListFile, VarListTab, VarOutPath, DataOutPath, ConvertDataOutPath, ExtractionSubPath, DemandDeliverySubPath, ModelSubPath, GroupDataDirPath, ScenarioDir, DVDssMin, DVDssMax, SVDssMin, SVDssMax, NameMin, NameMax, DirMin, DirMax, IndexMin, IndexMax, StartMin, StartMax, EndMin, EndMax, VarMin, VarMax, DemandFilePath, DemandFileName, DemandFileTab, DemMin, DemMax, InflowOutSubPath, InflowFilePath, InflowFileName, InflowFileTab, InflowMin, InflowMax

def convert_all_cfs_to_taf(df):
    """
    Convert all columns with units 'CFS' to 'TAF'
    Conversion: 1 CFS-month = 0.001984 * (days_in_month) TAF

    Parameters:
        df (pd.DataFrame): Main data DataFrame with MultiIndex columns.

    Returns:
        pd.DataFrame: DataFrame with converted columns (new columns labeled as 'TAF').
    """

    # Precompute days in each month
    days_in_month = df.index.days_in_month.to_numpy()

    columns_to_convert = []

    for col in df.columns:
        part_a, part_b, *_, data_unit = col

        if data_unit != "CFS":
            continue

        columns_to_convert.append(col)

    #print(f"\nConverting {len(columns_to_convert)} columns from CFS to TAF...")

    # Perform conversion
    for col in columns_to_convert:
        new_col = list(col)
        new_col[-1] = "TAF"
        new_col = tuple(new_col)
        df[new_col] = df[col].to_numpy() * 0.001984 * days_in_month
        #print(f"  ✓ {col} → {new_col}")

    return df
