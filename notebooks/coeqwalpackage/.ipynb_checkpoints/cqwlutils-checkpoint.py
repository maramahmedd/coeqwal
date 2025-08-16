import openpyxl
import re
import os

def get_xl_sheetnames(xlfn):
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    return(sheet_names)

def read_from_excel(xlfn, tabname, topleft, bottomright,hdr=True,dtypes=[]):
    # open excel file
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    inputs_sheet = wb.get_sheet_by_name(tabname)
    topsplit = re.split('(\d+)', topleft)
    botsplit = re.split('(\d+)', bottomright)
    
    if hdr:
        hdr_block = inputs_sheet[topleft:(botsplit[0]+topsplit[1])]
        hdr_values = [[str(v.value) for v in v1] for v1 in hdr_block][0]                        
        data_block = inputs_sheet[topsplit[0]+str(int(topsplit[1])+1):bottomright]
    else:
        data_block = inputs_sheet[topleft:bottomright]   
        hdr_values = None
    data_list = []
    # loop over months
    for i,row in enumerate(data_block):
        # loop over columns
        tmp  =[]
        if row[0].value=='null':
            pass
        else:
            if not dtypes: # list of dtypes is empty
                for j,cell in enumerate(row[:]):
                    tmp.append(str(cell.value))
            elif len(dtypes) != len(row): 
                print("Wrong number of dtypes provided - returning values as strings")
            else:
                for j,cell in enumerate(row[:]):
                    dtypi = dtypes[j]
                    if dtypi[0:2]=='dt':
                        parsetxt=dtypi[2:]
                        if type(cell.value) is not dt.datetime:  #check if it's already parsed as a datetime
                            tmp.append(dt.datetime.strptime(cell.value,parsetxt))
                        else:
                            tmp.append((cell.value))
                    elif dtypi[0:5]=='float':
                        tmp.append(float(cell.value))
                    else:
                        tmp.append(str(cell.value))
            data_list.append(tmp)
    return([hdr_values, data_list])


def read_init_file(CtrlFile, CtrlTab):

    # Hard-coded file structure
    ScenarioDirInd = 'B2'
    DssListFileInd = 'B3'
    DssListTabInd = 'B4'
    IndexMinInd = 'C5' # first scenario index
    IndexMaxInd = 'D5' # last scenario index
    ScenariosDirMinInd = 'C6' # first dss dir index
    ScenariosDirMaxInd = 'D6' # last dss dir index
    DssPathMinInd = 'C7' # first dss path index
    DssPathMaxInd = 'D7' # last dss path index
    StartMinInd = 'C8' # first start date index
    StartMaxInd = 'D8' # last start date index
    EndMinInd = 'C9' # First end date index
    EndMaxInd = 'D9' # Last end date index
    GroupDataDirInd = 'B10' # Group data extraction directory
    VarFileNameInd = 'B11' # Variable list file name
    VarFileTabInd = 'B12' # Variable list file tab
    VarMinInd = 'C13' # top left of variable name block
    VarMaxInd = 'D13' # bottom right of variable name block
    ExtractionDirInd = 'B14'
    ModelFilesDirInd = 'B15'
    ModelSubDirInd = 'B16'

    # Control File Example
    #Item	Name or description	Upper Left Cell	Lower Right Cell
    #Scenarios Directory	../../CalSim3_Model_Runs/Scenarios		
    #Scenario Listings File	coeqwal_cs3_scenario_listing_v2.xlsx		
    #Scenario Listings Tab	scenario_list		
    #Scenario Indices	Scenario identifiers	A1	A11
    #Scenario Directory Indices	Scenario directory names	C1	C11
    #DSS Path Indices	Dss path names	G1	G11
    #Start Date Indices	Start dates	H1	H11
    #Start Date Indices	End dates	I1	I11
    #Group Extraction Directory	Group_Data_Extraction		
    #Variables Listing File	trend_report_variables_v3.xlsx		
    #Variables List Tab	Variables List		E177
    #Variables List Indices	Variables List Block	D8	E177
    #Data Extraction Dir	Data_Extraction		
    #Model Files Dir	Model_Files		
    #Model Files SubDir	ModelFiles		

    # Read directory structure and contol file name
    Hdr, ScenarioDir = read_from_excel(CtrlFile, CtrlTab, ScenarioDirInd, ScenarioDirInd, hdr=False) # Scenarios directory (../../CalSim3_Model_Runs/Scenarios in the current structure)
    ScenarioDir = ScenarioDir[0][0]
    Hdr, DssListFile = read_from_excel(CtrlFile, CtrlTab, DssListFileInd, DssListFileInd, hdr=False) # DSS file names Excel workbook
    DssListFile = DssListFile[0][0]
    DssListPath = os.path.join(ScenarioDir,DssListFile) # path to DSS file names Excel workbook
    
    # Read file names and ranges for DSS and vars from control file
    Hdr, DssListTab = read_from_excel(CtrlFile, CtrlTab, DssListTabInd, DssListTabInd, hdr=False) # DSS file names Excel workbook Tab
    DssListTab = DssListTab[0][0]
    Hdr, IndexMin = read_from_excel(CtrlFile, CtrlTab, IndexMinInd, IndexMinInd, hdr=False) # Scenario Index Name UL
    IndexMin = IndexMin[0][0]
    Hdr, IndexMax = read_from_excel(CtrlFile, CtrlTab, IndexMaxInd, IndexMaxInd, hdr=False) # Scenario Index Name LR
    IndexMax = IndexMax[0][0]
    Hdr, DirMin = read_from_excel(CtrlFile, CtrlTab, ScenariosDirMinInd, ScenariosDirMinInd, hdr=False) # Scenario Dir Name UL
    DirMin = DirMin[0][0]
    Hdr, DirMax = read_from_excel(CtrlFile, CtrlTab, ScenariosDirMaxInd, ScenariosDirMaxInd, hdr=False) # Scenario Dir Name UL
    DirMax = DirMax[0][0]
    Hdr, DssMin = read_from_excel(CtrlFile, CtrlTab, DssPathMinInd, DssPathMinInd, hdr=False) # DSS Path Name UL
    DssMin = DssMin[0][0]
    Hdr, DssMax = read_from_excel(CtrlFile, CtrlTab, DssPathMaxInd, DssPathMaxInd, hdr=False) # DSS Path Name LR
    DssMax = DssMax[0][0]
    Hdr, StartMin = read_from_excel(CtrlFile, CtrlTab, StartMinInd, StartMinInd, hdr=False) # Start Date UL
    StartMin = StartMin[0][0]
    Hdr, StartMax = read_from_excel(CtrlFile, CtrlTab, StartMaxInd, StartMaxInd, hdr=False) # Start Date LR
    StartMax = StartMax[0][0]
    Hdr, EndMin = read_from_excel(CtrlFile, CtrlTab, EndMinInd, EndMinInd, hdr=False) # Start Date UL
    EndMin = EndMin[0][0]
    Hdr, EndMax = read_from_excel(CtrlFile, CtrlTab, EndMaxInd, EndMaxInd, hdr=False) # Start Date LR
    EndMax = EndMax[0][0]
    Hdr, GroupDataDirName = read_from_excel(CtrlFile, CtrlTab, GroupDataDirInd, GroupDataDirInd, hdr=False) # directory name for group data extraction (Group_Data_Extraction in current structure)
    GroupDataDirName = GroupDataDirName[0][0]
    GroupDataDirPath = os.path.join(ScenarioDir, GroupDataDirName) # group data extraction directory (../../CalSim3_Model_Runs/Scenarios/Group_Data_Extraction in the current structure)
    #print(GroupDataDirPath)
    Hdr, VarListFileName = read_from_excel(CtrlFile, CtrlTab, VarFileNameInd, VarFileNameInd, hdr=False) # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    VarListFileName = VarListFileName[0][0]
    Hdr, VarListTab = read_from_excel(CtrlFile, CtrlTab, VarFileTabInd, VarFileTabInd, hdr=False) # tab for variable listing (TrendReportVars_CS3 in current structure)
    VarListTab = VarListTab[0][0]
    Hdr, VarMin = read_from_excel(CtrlFile, CtrlTab, VarMinInd, VarMinInd, hdr=False) # variable listing UL
    VarMin = VarMin[0][0]
    Hdr, VarMax = read_from_excel(CtrlFile, CtrlTab, VarMaxInd, VarMaxInd, hdr=False) # variable listing LR
    VarMax = VarMax[0][0]
    Hdr, ExtractionDir = read_from_excel(CtrlFile, CtrlTab, ExtractionDirInd, ExtractionDirInd, hdr=False) #  Var extraction Dir Name
    ExtractionDir = ExtractionDir[0][0]
    Hdr, ModelFilesDir = read_from_excel(CtrlFile, CtrlTab, ModelFilesDirInd, ModelFilesDirInd, hdr=False) #  Var extraction Dir Name
    ModelFilesDir = ModelFilesDir[0][0]
    Hdr, ModelSubDir = read_from_excel(CtrlFile, CtrlTab, ModelSubDirInd, ModelSubDirInd, hdr=False) #  Var extraction SubDir Name
    ModelSubDir = ModelSubDir[0][0]

    # Construct file and directory names
    # File and directory names
    DssListFileCsv = DssListFile.replace(".xlsx", ".csv")
    DssNamesOut = 'DssNamesFrom_' + DssListFileCsv # output DSS names CSV
    DssIndicesOut = 'IndicesFrom_' + DssListFileCsv # output DSS indices CSV
    DssDirsOut = 'DirNamesFrom_' + DssListFileCsv # output directory names CSV
    DssNamesOutPath = os.path.join(GroupDataDirPath, DssNamesOut) # output DSS names CSV path
    DssIndicesOutPath = os.path.join(GroupDataDirPath, DssIndicesOut) # output DSS index names CSV path
    DssDirsOutPath = os.path.join(GroupDataDirPath, DssDirsOut) # output DSS dir names CSV path

    # list of relevant variables file, tab, and range (B & C parts)
    VarListName = os.path.splitext(VarListFileName)[0] # variable names file without extension
    VarListExt = os.path.splitext(VarListFileName)[1] # variable names file extension
    VarListFile = VarListName + VarListExt # full file name
    VarListFileCsv = VarListFile.replace(".xlsx", ".csv")
    VarListPath = os.path.join(ScenarioDir, VarListFile)
    VarOut = 'VarsFrom_' + VarListFileCsv # output compund variable names CSV
    VarOutPath = os.path.join(GroupDataDirPath, VarOut)
    DataOut = 'DataFrom_' + VarListFileCsv # file name for multi-study output CSV
    DataOutPath = os.path.join(GroupDataDirPath, DataOut) # file name for multi-study output CSV path
    ConvertDataOut = 'ConvertDataFrom_' + VarListFileCsv # file name for multi-study output CSV
    ConvertDataOutPath = os.path.join(GroupDataDirPath, ConvertDataOut) # file name for multi-study output CSV path
    ExtractionSubDir = 'Variables_From_' + VarListName + '_' + VarListTab
    ExtractionSubPath = os.path.join(ExtractionDir, ExtractionSubDir)
    ModelSubPath = os.path.join('Model_Files','DSS','output')

    # debug print
    # print(DssListFile)    
    # print(DssListTab)    
    # print(DssListPath)
    # print(DssNamesOutPath)
    # print(DssIndicesOutPath)
    # print(DssDirsOutPath)
    # print(VarListPath)
    # print(VarOutPath)
    # print(DataOutPath)
    # print(ExtractionSubPath)
    # print(ModelSubPath)
    # print(GroupDataDirPath)
 
    # return info
    return DssListFile, DssListTab, DssListPath, DssNamesOutPath, DssIndicesOutPath, DssDirsOutPath, VarListPath, VarListFile, VarListTab, VarOutPath, DataOutPath, ConvertDataOutPath, ExtractionSubPath, ModelSubPath, GroupDataDirPath, ScenarioDir, DssMin, DssMax, DirMin, DirMax, IndexMin, IndexMax, StartMin, StartMax, EndMin, EndMax, VarMin, VarMax