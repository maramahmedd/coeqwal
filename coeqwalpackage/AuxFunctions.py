# -*- coding: utf-8 -*-
"""
This file contains helper functions used in processing CalSim outputs
    Generally things like:
            - Exceendance calculations
            - Time series utilities (add WY, delivery year, etc)
            - others....
Putting these functions in a separate (this) file to hopefully make the main
file cleaner and easier to follow
          
Created on Wed Nov 29 13:35:27 2017

@author: jmgilbert
"""


#%% preliminaries

import pandas as pnd
import numpy as np
import openpyxl

import io # used in catalog condensing
import time

import re
import os
import datetime as dt

import dss3_functions_reference as dss  ####IMPORTANT - change import to reflect Python3 vs Python2.x versions!!!

#%% other utility functions - taken from PlotAllocDiff.py

def calc_exceed(series):
    #series = series.dropna()  # get rid of any missing values
    n = len(series)
    p_list = []
    for m in range(1,n+1):
        p_list.append(float(m)/(n+1))
    pc = zip(p_list, sorted(list(series), reverse=True))
    return(pc)

def calc_exceed_series(series, monfilt=0):
    series = series.dropna()  # get rid of any missing values
    if monfilt !=0: # filter out only values for a given month indicated by monfilt; monfilt==0 means all months
        dts_filt = [x[1] for x in zip(series.index, series) if x[0].month==monfilt]
    else:
        dts_filt = [x for x in series]

    pc = list(calc_exceed(dts_filt))
    probList = [i[0] for i in pc]
    sortedVals = [i[1] for i in pc]
    excdDF = pnd.Series(sortedVals, index=probList)
    return(excdDF)


def calc_exceed_df(df, var, wyt=None, monfilt=0):
    # returns a new dataframe with multiindex of study, exceedance prob
    createDF = True
    # if wyt argument is None, include all data, otherwise use
    # the non-zero positive value for selecting data for a particular WYT
    if wyt not in [None, 1,2,3,4,5]:
        print('Invalid water year type provided - try again!')
        exit
    
    # assumes study index is the first level of multi-index
    studies = df.index.levels[0]
    dts = df.index.levels[1]
    if monfilt !=0: # filter out only values for a given month indicated by monfilt; monfilt==0 means all months
        dts_filt = [x for x in dts if x.month==monfilt]
    else:
        dts_filt = [x for x in dts]
    for n,s in enumerate(studies): # loop through studies
        # assume var is a list with variables corresponding to each study, or a 
        # single value that is the variable name that is constant across studies
        if isinstance(var, list):
            varn = var[n]
        else:
            varn = var
        if wyt==None:
            tmpdf = df.loc[(s,dts_filt),varn]
        else:
            tmpdf = df[df['Q5_WYT']==wyt].loc[(s,),varn]
        if createDF:
            pc = calc_exceed(tmpdf)
            probList = [i[0] for i in pc]
            sortedVals = [i[1] for i in pc]
            mltidx = pnd.MultiIndex.from_product([list(studies), probList], names=['Study','ExcProb'])
            excdDF = pnd.DataFrame(index=mltidx)
            createDF=False
        else:
            pc = calc_exceed(tmpdf)
            probList = [i[0] for i in pc]
            sortedVals = [i[1] for i in pc]
        excdDF.loc[(s, ) , varn] = sortedVals
    return(excdDF)
            
def addDelYr(x):
    yr = x['Date'].year
    mo = x['Date'].month
    #yr = x.year
    #mo = x.month
    if mo < 2:
        dyr = yr-1
    else:
        dyr = yr
    return(dyr)
    
def addWY(x):
    #yr = x['Date'].year
    #mo = x['Date'].month
    yr = x.year
    mo = x.month
    if mo > 9:
        wyr = yr+1
    else:
        wyr = yr
    return(wyr)
    
def wymo(x):
    mo = x.month
    if mo > 9:
        wymnth = mo-9
    else:
        wymnth = mo+3
    return(wymnth)

def cfs_taf(x):
    #yr = x['Date'].year
    #mo = x['Date'].month
    dy = x['Date'].day
    v = (86400./43560.)*dy/1000.  
    return(v)

def taf_cfs(x):
    #yr = x['Date'].year
    #mo = x['Date'].month
    dy = x['Date'].day
    v = 1000/dy * (43560./86400.) #(86400./43560.)*dy/1000.  
    return(v)

def write_to_excel(xlfn, tabname, val_list, col_titles):
    #import openpyxl
     # write out to an excel file - one tab per demand unit?
    if not os.path.exists(xlfn):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(xlfn)
    #xlfn = os.path.join(os.path.dirname(idcoutputdss),'WestsideIDCOutputCompare.xlsx')
    
    startCol = 1 # column from which to start inserting data
    startRow = 1 # can adjust this - assumes you put header in 7 rows above this, so shouldn't be <7
    
    ws = wb.create_sheet(title=tabname)
    try:
        date_style = openpyxl.styles.Style(number_format="M/DD/YYY")
    except:
        print("Excel date style not working..trying something different")
        date_style = openpyxl.styles.numbers.FORMAT_DATE_XLSX14.replace('-','/')
    for i,c in enumerate(col_titles):
        _ = ws.cell(column=startCol+i, row=startRow, value=c)
    for i,r in enumerate(val_list):
        for j,v in enumerate(r):
            _ = ws.cell(column=startCol+j, row=startRow+i+1, value =v)
#            if j==0:
#                # assume this is a date column
#                _.style= date_style
    wb.save(filename=xlfn)

def get_xl_sheetnames(xlfn):
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    return(sheet_names)

def read_from_excel(xlfn, tabname, topleft, bottomright,hdr=True,dtypes=[]):
    # open excel file
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    inputs_sheet = wb.get_sheet_by_name(tabname)
    topsplit = re.split('(\d+)', topleft)
    botsplit = re.split('(\d+)', bottomright)
    
    if hdr:
        hdr_block = inputs_sheet[topleft:(botsplit[0]+topsplit[1])]
        hdr_values = [[str(v.value) for v in v1] for v1 in hdr_block][0]                        
        data_block = inputs_sheet[topsplit[0]+str(int(topsplit[1])+1):bottomright]
    else:
        data_block = inputs_sheet[topleft:bottomright]   
        hdr_values = None
    data_list = []
    # loop over months
    for i,row in enumerate(data_block):
        # loop over columns
        tmp  =[]
        if row[0].value=='null':
            pass
        else:
            if not dtypes: # list of dtypes is empty
                for j,cell in enumerate(row[:]):
                    tmp.append(str(cell.value))
            elif len(dtypes) != len(row): 
                print("Wrong number of dtypes provided - returning values as strings")
            else:
                for j,cell in enumerate(row[:]):
                    dtypi = dtypes[j]
                    if dtypi[0:2]=='dt':
                        parsetxt=dtypi[2:]
                        if type(cell.value) is not dt.datetime:  #check if it's already parsed as a datetime
                            tmp.append(dt.datetime.strptime(cell.value,parsetxt))
                        else:
                            tmp.append((cell.value))
                    elif dtypi[0:5]=='float':
                        tmp.append(float(cell.value))
                    else:
                        tmp.append(str(cell.value))
            data_list.append(tmp)
    return([hdr_values, data_list])

    
def readVarList(varListfp, plotGroup=False):
    # revised version that expects the first line to contain the number (N) of 
    # study files to compare; then assumes next N lines are the paths to those
    # files; then reads remaining lines as the variables to extract
    # First column in file path line is the study nickname (short) if provided; 
    # second column is a longer description; third column is whether this path
    # refers to a CalSimII (=2) or CalSim3 (=3) study; fourth column is file path
    # In variable lines - 2 scenarios:
    # 1)If all studies are of same CalSim type (i.e. all CalSimII or all CalSim3):
    #   First column is a category value (number or letter - every line with
    # the same category value will be grouped in output)
    #   Second column is variable description (optional)
    #    Third column is variable path in dss file
    # 2) If some studies are CSII and some are CS3:
    #    repeat pattern of scenario 1 (3 columns at a time) for each study 
    #    listed in the study rows
    with open(varListfp, 'r') as vl:
        vLines = vl.readlines()
         
    numFiles = int(vLines[0].split(',')[0].strip())
    print("Comparing %s studies" %numFiles)
    
    studyDict = {}
    print("Those studies are:\n")
    studyTypes = []
    for i in range(int(numFiles)):
        ls = vLines[i+1].split(',')  # assumes a comma separated file
        print("      %s" %ls[3])
        studyTypes.append(ls[2].strip())
        # studyDict structure is:
        # key = integer index starting at 0
        # values = list of the following variables:
        #           0  - study short name
        #           1 - study long name/description
        #           2 - study type (2 = CalSimII, 3 = CalSim3)
        #           3 - path to file
        #           4 - slot for DSS reading IFLTABI variable (initialized as None)
        #           5 - slot for a list of vars/paths (initialized as None)
        #           6 - slot for catalog status (is it open? how many records? could add date/time for updates, etc)
        #           7 - slot for the complete DSS catalog (initialied as None)
        #           8 - slot for a list of categories/groups of variables to summarize on
        #studyDict[ls[3].strip()] = [ls[0].strip(), ls[1].strip(), ls[2].strip()] 
        studyDict[i] = [ls[0].strip(), ls[1].strip(), ls[2].strip(), ls[3].strip(), None, None, None, None, None] 


    
#    studyTypes = [studyDict[i][2] for i in studyDict.keys()]
#    if len(list(set(studyTypes)))>1:   # assume that if not all same type (i.e., all CSII or all CS3)
#        numVarSets = len(studyTypes)   # then number of variable sets is the number of studies 
#    else:
#        numVarSets = 1
    numVarSets = 1   # assume the number of sets of variable definitions is 1 to begin with
    if len(set(studyTypes)) > 1:
        numVarSets = numFiles   # if you're mixing CSII and CS3, then you'll 
                                # need to read in a separate variable definition 
                                # set for each indiviiaul file
    
    # we want to check for a blank row so we stop reading any comments or extraneous
    # text below it - a blank row will be a series of ',' of length (numVarSets*3)+1 if numVarSets>1; and 3 if numVarSets=1     
    tmp_str = ''
    if numVarSets==1:
        for n in range(3):
            tmp_str+=','
        tmp_str+='\n'
    else:
        for n in range(numVarSets*3 +1):
            tmp_str+=','
        tmp_str+='\n'
    
    for n,v in enumerate(vLines):
        if v==tmp_str:
            end_file = n
            break
        else:
            end_file=len(vLines)
    vLines = vLines[0:end_file]
    # now read in variable names along with category and descriptive names, if
    # present
    for v in range(numVarSets):
        vars = {}
        cats = [] # list to collect the categories in
        for l in vLines[numFiles+1:]:
            
            ls = l.split(',')[(v*5):(v*5)+4]
            #print(ls)
            if ls[0]!='':
                categ = int(ls[0].strip())
                descrip = ls[1].strip()
                vari = ls[2].strip()
                if plotGroup:
                    plotGrp = ls[3].strip()
                    vars[vari] = [categ, descrip, plotGrp]
                else:
                    vars[vari] = [categ, descrip]
                cats.append(categ)

        uniq_cats = list(set(cats))
        uniq_cat_dict = {}
        for u in uniq_cats:
            tmp=[]
            for vi in vars:
                if vars[vi][0]==u:
                    tmp.append(vi)
            uniq_cat_dict[u]=tmp
        if numVarSets==1:
            for k in studyDict.keys():
                studyDict[k][5] = vars
                studyDict[k][8] = uniq_cat_dict
        else:
            studyDict[v][5] = vars
            studyDict[v][8] = uniq_cat_dict
    
    return(studyDict)
    
def condense_path(pl):
    # takes a pathlist and condenses the time dimension
    path_dict = {}
    Dlist =[]
    cntr = 0
    for n,r in enumerate(pl):
        rs = r.split("/")
        try:
            if rs[-1].strip()=='':
                [ldngspace, Apt,Bpt,Cpt,Dpt,Ept,Fpt, end] = r.split("/")
            else:
                [ldngspace, Apt,Bpt,Cpt,Dpt,Ept,Fpt] = r.split("/")
        except:
            print("Didn't read in this variable correctly: %s - %s  - %s" %(n,r, len(r.split("/"))))
        if n==len(pl)-1: #then at the end of list - can't check next
            #print("should just be seeing this once")
           # do a backwards check to see if it's different than the previous
            prevBCpt = pl[n-1].split("/")[2]+"-"+pl[n-1].split("/")[3]
            if (Bpt+"-"+Cpt)==prevBCpt:
# at the last in this variable time-group
                Dlist.append(Dpt)
                Drange=[Dlist[0], Dlist[-1]]
                Dlist = []
                path_dict[cntr] = {"A":Apt,"B":Bpt,"C":Cpt,"D":Drange, "E":Ept,"F":Fpt}
            else:
                path_dict[cntr] = {"A":Apt,"B":Bpt,"C":Cpt,"D":Dpt, "E":Ept,"F":Fpt}
                cntr+=1
        else:
            nextBCpt = pl[n+1].split("/")[2]+"-"+pl[n+1].split("/")[3]
            if nextBCpt==(Bpt+"-"+Cpt):
                Dlist.append(Dpt)
            else:
                # at the last in this variable time-group
                Dlist.append(Dpt)
                Drange=[Dlist[0], Dlist[-1]]
                Dlist = []
                path_dict[cntr] = {"A":Apt,"B":Bpt,"C":Cpt,"D":Drange, "E":Ept,"F":Fpt}
                cntr+=1
    #jsontt= json.loads(json.dumps((path_dict)))
    #json_sort = sorted(jsontt, key=jsontt.keys())
    return(path_dict)
    
    
    
def getValAtTime(studyDict, startDate, ntimes=1):
    '''
    Function to read value for many variables at a specified time
    Input:  studyDict  = dictionary containing list of variables to lookup
            selectDate = date at which to select/extract data
    Returns:   dictionary of variable-values
    '''
    cdate = dt.datetime.strftime(startDate, '%d%b%Y') #takes a python datetime object and converts to string format ->eg '31Oct1921'
    ctime = '2400'
    data = {} # data will be returned in a dictionary with keys=variable name, value=value at the time of interest
    for s in studyDict:
        [ifltab, iostat] = dss.open_dss(studyDict[s][3])
        if iostat==0:
            print("Opened file: %s"  %studyDict[s][3])
            studyDict[s][4] = ifltab
        # get the condensed path list for this study
        #cpl = [studyDict[s][5][v][2] for v in studyDict[s][5]]
        # loop through variable dictionary (assumes it includes the condensed list as the 3rd item [index 2] in value list for each key)
        for var in studyDict[s][5]:
            cpath=studyDict[s][5][var][2]
            [nvals, vals, cunits, ctype, iofset, istat] = dss.read_regts(ifltab, cpath, cdate, ctime, ntimes)
            for i,v in enumerate(vals):
                if v==-902.0 or v==-901.0:
                    vals[i] = np.NaN
#            if cunits.value.strip().upper() =='CFS':
#                vals = [v[0]*v[1] for v in zip(vals,cfs_taf_v)]
#            else:
            vals = [v for v in vals]
            #data[s].loc[var] = vals
            data[var] = vals[0]
            #data[s].loc[(data[s][var]==-902.0),var] = np.NaN
            #data[s].loc[(data[s][var]==-901.0),var] = np.NaN
        ret = dss.close_dss(ifltab)  #close it down!
        studyDict[s][4] = None # set ifltab variable back to None
        print(ret) 
    
    return(data)
    
def perToDTlist(periodidx,include_time=False):
    dtlist=[]
    for d in list(periodidx):
        y = d.year
        m = d.month
        dy = d.day
        if include_time:
            h = d.hour
            M = d.minute
            odt = dt.datetime(y, m, dy, h, M)
        else:
            odt = dt.date(y, m , dy)
        dtlist.append(odt)
    return(dtlist)

def condense_cat(pathlist, drop_parts=['D'], group_parts=['B']):
    t1 = time.time()
    dftest = pnd.read_table(io.StringIO('\n'.join(pathlist)), delimiter='/',
                            names=['Blank1', 'A','B','C','D','E','F', 'Blank2'],
                            keep_default_na=False)
    t2 = time.time()
    print("Took %s sec to convert paths to dataframe" %(t2-t1))
    dftest = dftest.drop(['Blank1', 'Blank2'], axis=1)
    all_parts = ['A', 'B', 'C', 'D', 'E', 'F']
    if len(drop_parts)==0:
        print('No drop parts provided - the returned catalog\npath list will be the same as you started with')
        keep_parts = all_parts
        #dftest['full_path'] = ['/'+c[1].A + '/' + c[1].B + '/' + c[1].C + '/' + c[1].D + '/' + c[1].E + '/'+ c[1].F +'/' for c in dftest.iterrows()]
    else:
        keep_parts = [x if x not in drop_parts else '' for x in all_parts]
        
        #dftest['full_path'] = ['/'+'/'.join([c[1][k] if k !='' else '' for k in keep_parts])+'/' for c in dftest.iterrows()]
    t3 = time.time()
    #print("Took %s sec to rebuild modified path" %(t3-t2))
    condcat = dftest.groupby(group_parts).nth(0) #keep the first value in each group
    t4 = time.time()
    print("Took %s sec to do group-by operation" %(t4-t3))
    #this is what I had originally, modified for HMS DSS file work--> condcat['search_col'] = ['/'.join(c) if type(c)==tuple else c for c in condcat.index]
    condcat['search_col'] = ['/'.join([str(c1) for c1 in c]) if type(c)==tuple else c for c in condcat.index]
    condcat.reset_index(inplace=True)
    t5 = time.time()
    print("Took %s sec to add column" %(t5-t4))
    
    condcat['full_path'] = ['/'+'/'.join([str(c[1][k]) if k !='' else '' for k in keep_parts])+'/' for c in condcat.iterrows()]
    t6 = time.time()
    print("Took %s sec to rebuild modified path" %(t6-t5))
    return(condcat)

def select_paths(condcat, study):
    '''
        a function that takes a condensed list of ALL catalog paths (as a dataframe) and an arbitrary
        list of variables/parts to retrieve; returns list of DSS-ready (no D-part)
        paths for direct use in DSS-data retrieval
    '''
    if type(condcat)!=pnd.core.frame.DataFrame:
        print("The `condcat` variable does not appear to be a dataframe - try again!")
        return()
    
    if not 'search_col' in condcat.columns:
        print("No `search_col` column in dataframe - try running through the\ncondense_cat function first.")
        return()
        
    listoB = []
    for i in list(study[5].keys()):
        if '/' in i: # assume this means it's a part or whole path
            ls = i.split('/')
            if i[0]=='/':
                if[-1]=='/':
                    lsi = ls[1:-1]
                else:
                    lsi = ls[1:]
            else:
                if[-1]=='/':
                    lsi = ls[0:-1]
                else:
                    lsi = ls[0:]
            listoB.append(lsi[1]+'/'+lsi[2])
        else:
            listoB.append(i)
    #listoB = [i.split('/') for i in studyDict[s][5].keys()] 
    sel_condcat = condcat[condcat.search_col.isin(listoB)]
    sel_condcat.reset_index(inplace=True)
    #sel_condcat.set_index('full_path', inplace=True)
    
    vari = study[5].keys()
    for i in vari:
        #pts = sel_condcat.loc[i]
        cpath = sel_condcat[sel_condcat['search_col']==i]['full_path'].iloc[0] #'/'+ '/'.join([pts['A'],pts['B'],pts['C'],'',pts['E'],pts['F']])+'/'  #pts['B']

        study[5][i].append(cpath)
    
    return(study)
        
def get_catalogs(studyDict, condense_parts=['B','C']):
    '''
        a convenience function to get catalog entries; does the following:
            - gets all paths in file
            - does path condensing (removes D-parts)
    '''
    catunit = 13
    for s in studyDict.keys():
        fp  = studyDict[s][3]  # get path of DSS file
        dss.fortran_close_file(catunit)  # make sure it's closed before we try to open it
        #[lgenca, lopnca, lcatlg, lgencd, lopncd, lcatcd, nrecs] = dss.open_catalog('', 12)
        if studyDict[s][6]==None or studyDict[s][6]['LOPNCA']==False:

            [lgenca, lopnca, lcatlg, lgencd, lopncd, lcatcd, nrecs] = dss.open_catalog(fp, catunit)

        else:
            print("Didn't open catalog file")
            nrecs=0
        print(nrecs)
        cat_dict = {}
        cat_dict['LOPNCA'] = lopnca
        cat_dict['NRECS'] = nrecs
        studyDict[s][6]=cat_dict
        [studyDict[s][7], lopnca] = dss.read_catalog(lopnca, icunitin=catunit)
        if len(studyDict[s][7])==0:
            print("didn't retrieve the catalog")
            dss.fortran_close_file(catunit)  # close it again, just to be safe
            studyDict[s][6]['LOPNCA'] = False
            return(studyDict)
        studyDict[s][6]['LOPNCA'] = lopnca
        dss.fortran_close_file(catunit)  # close it again, just to be safe
        
        dftest = pnd.read_table(io.StringIO('\n'.join(studyDict[s][7])), delimiter='/',names=['Blank1', 'A','B','C','D','E','F', 'Blank2'])
        dftest = dftest.drop(['Blank1', 'Blank2'], axis=1)
        #dftest['full_path'] = pnd.read_table(io.StringIO('\n'.join(studyDict[s][7])),names=['full_path'])
        dftest['full_path'] = ['/'+c[1].A + '/' + c[1].B + '/' + c[1].C + '//' + c[1].E + '/'+ c[1].F +'/' for c in dftest.iterrows()]
    
        #condcat = dftest.groupby(['B','C']).nth(0) #keep the first value in each group
        #condcat['search_col'] = [c[0]+'/'+c[1] for c in condcat.index]
        condcat = dftest.groupby(['B']).nth(0) #keep the first value in each group
        condcat['search_col'] = [c for c in condcat.index]
    
        listoB = []
        for i in studyDict[s][5].keys():
            if '/' in i: # assume this means it's a part or whole path
                ls = i.split('/')
                if i[0]=='/':
                    if[-1]=='/':
                        lsi = ls[1:-1]
                    else:
                        lsi = ls[1:]
                else:
                    if[-1]=='/':
                        lsi = ls[0:-1]
                    else:
                        lsi = ls[0:]
                listoB.append(lsi[1]+'/'+lsi[2])
            else:
                listoB.append(i)
        #listoB = [i.split('/') for i in studyDict[s][5].keys()] 
        sel_condcat = condcat[condcat.search_col.isin(listoB)]
        sel_condcat.reset_index(inplace=True)
        #sel_condcat.set_index('full_path', inplace=True)
        sel_condcat.set_index('search_col', inplace=True)
        
        vari = studyDict[s][5].keys()
        try:
            for i in vari:
                pts = sel_condcat.loc[i]
                cpath = '/'+ '/'.join([pts['A'],pts['B'],pts['C'],'',pts['E'],pts['F']])+'/'  #pts['B']
                studyDict[s][5][i].append(cpath)
        except:
            print("didn't work...returning sel_condcat")
            return([condcat, sel_condcat])
    
    return(studyDict)
    
def get_dss_data(studyDict, data, cdate, ctime, ntimes, cfs_taf_v, col_name_parts=[2,3], unit_convert=False):
    ''' 
        another convenience function - get's regular time series data from dss files
        data is a pandas dataframe that's been set up with an appropriate datetime index already
    '''
    units={}
    # open the files, add ifltab to dict
    for s in studyDict:
        thisStudyUnits = {}  #let's keep track of units
        [ifltab, iostat] = dss.open_dss(studyDict[s][3])
        if iostat==0:
            print("Opened file: %s"  %studyDict[s][3])
            studyDict[s][4] = ifltab
        # get the condensed path list for this study
        #cpl = [studyDict[s][5][v][2] for v in studyDict[s][5]]
        cpl = [v for v in studyDict[s][5]]
        for cpath in cpl:
            varpts = [cpath.split('/')[k] for k in col_name_parts]
            var = '_'.join(varpts)
            [nvals, vals, cunits, ctype, iofset, istat] = dss.read_regts(ifltab, cpath, cdate, ctime, ntimes)
            if istat!=0:
                print("Error retrieving DSS time series for path %s\nistat=%s" %(cpath,istat))
                
            for i,v in enumerate(vals):
                if v==-902.0 or v==-901.0:
                    vals[i] = np.NaN
            if cunits.strip().upper() =='CFS' and unit_convert:
                vals = [v[0]*v[1] for v in zip(vals,cfs_taf_v)]  # convert to TAF
                thisUnits = 'TAF'
            else:   
                vals = [v for v in vals]
                thisUnits = cunits.strip().upper()
            thisStudyUnits[var] = thisUnits
            #data[s].loc[var] = vals
            #data[s][var] = vals

            data.loc[(s),var] = vals
            #data.loc[(s),var]
            data[data[var]==-902.0] = np.NaN #.loc[(s,var==-902.0),var] = np.NaN
            data[data[var]==-901.0] = np.NaN
            #data.loc[(s,data[s][var]==-902.0),var] = np.NaN
            #data[s].loc[(data[s][var]==-901.0),var] = np.NaN
    
        units[s] = thisStudyUnits
        #studyDict[s]['Units']=thisStudyUnits
        ret = dss.close_dss(ifltab)  #close it down!
        studyDict[s][4] = None # set ifltab variable back to None
        print(ret)
        
    return([data, units])